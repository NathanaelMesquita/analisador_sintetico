import streamlit as st
import pandas as pd
import os
import openpyxl 
import re
from openpyxl import load_workbook
from io import BytesIO

# Configurações iniciais da página
st.set_page_config(page_title="Análise Sintética de RIF", layout="wide")

# Caminho relativo para os logotipos
caminho_base = os.path.dirname(__file__)
logo_esquerda = os.path.join(caminho_base, "logo_esquerda.png")
logo_direita = os.path.join(caminho_base, "logo_direita.png")

# Adiciona logotipo redimensionado e posiciona acima do título
col1, _, col2 = st.columns([1, 5, 1])

with col1:
    st.image(logo_esquerda, width=100)

with col2:
    st.image(logo_direita, width=100)

st.title("Análise Sintética de RIF")

# Carregar arquivos de entrada
principais_envolvidos = st.file_uploader("Carregue o arquivo 'Principais Envolvidos.xlsx'", type="xlsx")
informacoes_adicionais = st.file_uploader("Carregue o arquivo 'InformacoesAdicionais.xlsx'", type="xlsx")

# Aviso sobre o local do arquivo de saída
st.info("Obs.: Será gerado um arquivo chamado análises_sintéticas.xlsx na mesma pasta onde o script está localizado.")

# Caminho do arquivo de saída
output_file = os.path.join(os.getcwd(), "análises_sintéticas.xlsx")

# Botão para iniciar a análise
if st.button("Gerar Análise"):
    if principais_envolvidos is not None and informacoes_adicionais is not None:
        # Carregar planilhas em DataFrames
        principais_envolvidos_df = pd.read_excel(principais_envolvidos, engine="openpyxl", usecols="B")
        informacoes_adicionais_df = pd.read_excel(informacoes_adicionais, engine="openpyxl")
        envolvidos = principais_envolvidos_df.iloc[:, 0].dropna().tolist()  # Lista de envolvidos

        # Criar lista para armazenar os dados a serem escritos
        all_data_written = False

        # Verifica se o arquivo já existe
        if os.path.exists(output_file):
            # Se o arquivo já existir, abre em modo append
            mode = "a"
        else:
            # Se o arquivo não existir, cria um novo
            mode = "w"

        # Processar dados e gravar no arquivo
        with pd.ExcelWriter(output_file, engine="openpyxl", mode=mode) as writer:
            for envolvido_cpf_cnpj in envolvidos:
                safe_cpf_cnpj = re.sub(r'[\/:*?"<>|]', "_", str(envolvido_cpf_cnpj))
                remetente_sheet_name = f"{safe_cpf_cnpj}_REMETENTE"
                beneficiario_sheet_name = f"{safe_cpf_cnpj}_BENEFICIARIO"

                # Filtrar dados de REMETENTES
                remetentes = informacoes_adicionais_df[
                    (informacoes_adicionais_df["REMETENTE/BENEFICIARIO CPF/CNPJ"] == envolvido_cpf_cnpj) &
                    (informacoes_adicionais_df["REMETENTE OU BENEFICIARIO?"] == "REMETENTE")
                ][[
                    "RIF", "REMETENTE/BENEFICIARIO CPF/CNPJ", "REMETENTE/BENEFICIARIO NOME",
                    "REMETENTE OU BENEFICIARIO?", "VALOR", "TITULAR CPF/CNPJ",
                    "TITULAR NOME", "DATA/PERÍODO"
                ]]

                if not remetentes.empty:
                    remetentes.to_excel(writer, sheet_name=remetente_sheet_name, index=False)
                    all_data_written = True  # Indicador de que dados foram escritos

                # Filtrar dados de BENEFICIÁRIOS
                beneficiarios = informacoes_adicionais_df[
                    (informacoes_adicionais_df["REMETENTE/BENEFICIARIO CPF/CNPJ"] == envolvido_cpf_cnpj) &
                    (informacoes_adicionais_df["REMETENTE OU BENEFICIARIO?"] == "BENEFICIARIO")
                ][[
                    "RIF", "REMETENTE/BENEFICIARIO CPF/CNPJ", "REMETENTE/BENEFICIARIO NOME",
                    "REMETENTE OU BENEFICIARIO?", "VALOR", "TITULAR CPF/CNPJ",
                    "TITULAR NOME", "DATA/PERÍODO"
                ]]

                if not beneficiarios.empty:
                    beneficiarios.to_excel(writer, sheet_name=beneficiario_sheet_name, index=False)
                    all_data_written = True  # Indicador de que dados foram escritos

        # Verificar se dados foram escritos
        if not all_data_written:
            st.warning("Nenhum dado foi escrito no arquivo. Verifique se os arquivos de entrada estão corretos.")
        else:
            # Ajuste de largura de colunas e formatação de moeda
            wb = load_workbook(output_file)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for column_cells in sheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.1
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Formatação de moeda para a coluna "VALOR"
                for cell in sheet["E"]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'R$ #,##0.00'

            # Salvar o arquivo com ajustes finais
            wb.save(output_file)
            st.success(f"Análise gerada com sucesso! Arquivo salvo em: {output_file}")

            # Botão para download do arquivo gerado
            with open(output_file, "rb") as f:
                st.download_button(
                    label="Baixar Relatório de Análise Sintética",
                    data=f,
                    file_name="análises_sintéticas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    else:
        st.error("Por favor, carregue os arquivos necessários antes de gerar a análise.")
